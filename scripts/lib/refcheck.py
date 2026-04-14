#!/usr/bin/env python3
"""uarch-bench 参考值管理与结果比对工具。

用法:
  refcheck.py update [--xlsx PATH] [--out PATH]          从 Excel 提取参考值
  refcheck.py check  <results.csv> [--ref PATH] [--tol FLOAT]  比对结果
"""

import json, csv, sys, os, re, argparse
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent.parent
PROJECT_DIR = SCRIPT_DIR.parent
DEFAULT_XLSX = PROJECT_DIR / "doc" / "RVA23_Instruction_Classification.xlsx"
DEFAULT_REF  = PROJECT_DIR / "data" / "reference.json"

# ── 颜色 ────────────────────────────────────────────────────────────
USE_COLOR = sys.stdout.isatty()
def _c(code, s): return f"\033[{code}m{s}\033[0m" if USE_COLOR else s
def green(s):  return _c("32", s)
def red(s):    return _c("31", s)
def yellow(s): return _c("33", s)
def cyan(s):   return _c("36", s)
def bold(s):   return _c("1", s)
def dim(s):    return _c("2", s)


# ── Excel → 参考 JSON ───────────────────────────────────────────────

def parse_excel(xlsx_path):
    """从 Excel 提取 {MNEMONIC: {category, latency, throughput}}。"""
    try:
        import openpyxl
    except ImportError:
        print("错误: 未安装 openpyxl，请执行: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["Instruction Classification"]

    ref = {}
    for r in range(2, ws.max_row + 1):
        cat  = ws.cell(r, 4).value
        mn   = ws.cell(r, 6).value
        lat  = ws.cell(r, 15).value
        thr  = ws.cell(r, 16).value

        if not mn or not cat:
            continue

        mn  = mn.strip().upper()
        entry = {"category": cat.strip()}

        if lat is not None:
            entry["latency"] = str(lat).strip()
        if thr is not None:
            entry["throughput"] = str(thr).strip()

        ref[mn] = entry

    return ref


def save_reference(ref, json_path):
    json_path = Path(json_path)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w") as f:
        json.dump(ref, f, indent=2, ensure_ascii=False)
    return json_path


def load_reference(json_path):
    with open(json_path) as f:
        return json.load(f)


# ── 数值解析 ─────────────────────────────────────────────────────────

def parse_numeric(value_str):
    """将 Excel 中可能复杂的值解析为浮点数。

    支持格式:
      "1"             → 1.0
      "17-18"         → 17.0  (区间 → 取下界)
      "1/17"          → 1/17  (分数)
      "2 (L1 hit)..." → 2.0  (取第一个数字)
      "—" / None      → None
    """
    if value_str is None:
        return None
    s = str(value_str).strip()
    if not s or s in ("—", "-", "N/A", ""):
        return None

    # 分数，如 "1/17"
    m = re.match(r"^(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)$", s)
    if m:
        denom = float(m.group(2))
        return float(m.group(1)) / denom if denom else None

    # 区间，如 "17-18" 或 "17~18"
    m = re.match(r"^(\d+(?:\.\d+)?)\s*[-~]\s*\d+", s)
    if m:
        return float(m.group(1))

    # 复杂字符串中的第一个数字
    m = re.match(r"(\d+(?:\.\d+)?)", s)
    if m:
        return float(m.group(1))

    return None


def extract_mnemonic(variant):
    """'ADD lat' → 'ADD', 'FCLASS.D tput' → 'FCLASS.D'"""
    for suffix in (" lat", " tput"):
        if variant.endswith(suffix):
            return variant[: -len(suffix)].strip().upper()
    return variant.strip().upper()


# ── 比对逻辑 ─────────────────────────────────────────────────────────

def check_results(csv_path, ref):
    """比对 CSV 结果与参考值（四舍五入后对比）。

    返回 (matches, mismatches, skipped)。
    """
    matches    = []
    mismatches = []
    skipped    = []

    with open(csv_path) as f:
        reader = csv.DictReader(f)
        for row in reader:
            test    = row["test"]
            variant = row["variant"]
            cpi     = float(row["cycles_per_iter"])
            ipi     = int(row["instr_per_iter"])

            if test.startswith("latency/"):
                test_type = "latency"
            elif test.startswith("throughput/"):
                test_type = "throughput"
            else:
                continue

            mnemonic = extract_mnemonic(variant)

            if mnemonic not in ref:
                skipped.append({
                    "mnemonic": mnemonic, "test": test, "variant": variant,
                    "type": test_type, "reason": "EXCEL_中无此指令",
                })
                continue

            entry = ref[mnemonic]

            if test_type == "latency":
                expected = parse_numeric(entry.get("latency"))
                if expected is None:
                    skipped.append({
                        "mnemonic": mnemonic, "test": test, "variant": variant,
                        "type": "latency", "reason": "EXCEL_无_latency_值",
                    })
                    continue

                rounded = round(cpi)
                ok = (rounded == int(expected))
                item = {
                    "mnemonic": mnemonic, "test": test, "variant": variant,
                    "type": "latency",
                    "measured_raw": round(cpi, 2),
                    "measured": rounded,
                    "expected": int(expected),
                    "expected_raw": entry.get("latency", ""),
                    "insn_per_iter": ipi,
                }
                (matches if ok else mismatches).append(item)

            elif test_type == "throughput":
                expected_thr = parse_numeric(entry.get("throughput"))
                if expected_thr is None or expected_thr == 0:
                    skipped.append({
                        "mnemonic": mnemonic, "test": test, "variant": variant,
                        "type": "throughput", "reason": "EXCEL_无_throughput_值",
                    })
                    continue

                measured_thr = 1.0 / cpi if cpi > 0 else 0
                rounded = round(measured_thr)
                expected_int = round(expected_thr)
                ok = (rounded == expected_int)
                item = {
                    "mnemonic": mnemonic, "test": test, "variant": variant,
                    "type": "throughput",
                    "measured_cpi": round(cpi, 4),
                    "measured_thr_raw": round(measured_thr, 2),
                    "measured_thr": rounded,
                    "expected_thr": expected_int,
                    "expected_raw": entry.get("throughput", ""),
                }
                (matches if ok else mismatches).append(item)

    return matches, mismatches, skipped


# ── 报告输出 ─────────────────────────────────────────────────────────

def print_report(matches, mismatches, skipped):
    total_checked = len(matches) + len(mismatches)

    # ── 不一致项 ──
    if mismatches:
        print(bold(red("\n" + "=" * 70)))
        print(bold(red("  不一致项: %d" % len(mismatches))))
        print(bold(red("=" * 70)))

        lat_mm = [m for m in mismatches if m["type"] == "latency"]
        thr_mm = [m for m in mismatches if m["type"] == "throughput"]

        if lat_mm:
            print(bold("\n  Latency 不一致:"))
            hdr = "  %-18s %10s %10s %14s  %s" % (
                "指令", "实测(四舍五入)", "期望", "Excel原值", "测试路径")
            print(dim(hdr))
            print(dim("  %s   %s   %s   %s  %s" % ("-"*16, "-"*8, "-"*8, "-"*12, "-"*20)))
            for m in sorted(lat_mm, key=lambda x: x["mnemonic"]):
                note = ""
                ipi = m.get("insn_per_iter", 1)
                if ipi > 1:
                    note = dim("  (往返链, ipi=%d)" % ipi)
                mn = red("%-18s" % m["mnemonic"])
                raw = m["expected_raw"]
                tst = dim(m["test"])
                print("  %s %5d(%5.2f) %10d %14s  %s%s" % (
                    mn, m["measured"], m["measured_raw"],
                    m["expected"], raw, tst, note))

        if thr_mm:
            print(bold("\n  Throughput 不一致:"))
            hdr = "  %-18s %10s %10s %14s  %s" % (
                "指令", "实测(四舍五入)", "期望", "Excel原值", "测试路径")
            print(dim(hdr))
            print(dim("  %s   %s   %s   %s  %s" % (
                "-"*16, "-"*8, "-"*8, "-"*12, "-"*20)))
            for m in sorted(thr_mm, key=lambda x: x["mnemonic"]):
                mn = red("%-18s" % m["mnemonic"])
                raw = m["expected_raw"]
                tst = dim(m["test"])
                print("  %s %5d(%5.2f) %10d %14s  %s" % (
                    mn, m["measured_thr"], m["measured_thr_raw"],
                    m["expected_thr"], raw, tst))

    if skipped:
        print(bold(yellow("\n" + "=" * 70)))
        print(bold(yellow("  跳过明细: %d" % len(skipped))))
        print(bold(yellow("=" * 70)))
        by_reason = {}
        for s in skipped:
            by_reason.setdefault(s["reason"], []).append(s)
        for reason, items in sorted(by_reason.items()):
            print(dim("    %s: %d 条指令" % (reason, len(items))))
            for s in sorted(items, key=lambda x: (x["test"], x["variant"])):
                print(dim("      %-16s %-10s %s :: %s" % (
                    s["mnemonic"],
                    s["type"],
                    s["test"],
                    s["variant"],
                )))

    # ── 汇总 ──
    print("\n" + "=" * 70)
    summary = "  %s: %d   %s: %d   %s: %d   共检查: %d" % (
        green("通过"), len(matches),
        red("不一致"), len(mismatches),
        yellow("跳过"), len(skipped),
        total_checked)
    print(summary)
    print("=" * 70)

    if not mismatches:
        print(green("\n  所有检查项均与 Excel 参考值一致。"))

    return len(mismatches) == 0


# ── CLI ──────────────────────────────────────────────────────────────

def cmd_update(args):
    xlsx = Path(args.xlsx)
    out  = Path(args.out)
    if not xlsx.exists():
        print("错误: 找不到 Excel 文件: %s" % xlsx, file=sys.stderr)
        sys.exit(1)

    print("读取: %s" % xlsx)
    ref = parse_excel(xlsx)
    save_reference(ref, out)
    print("写入: %s  (%d 条指令)" % (out, len(ref)))

    # 按分类汇总
    cats = {}
    for mn, e in ref.items():
        c = e["category"]
        cats[c] = cats.get(c, 0) + 1
    print("\n分类统计: %d 类" % len(cats))
    for c in sorted(cats):
        print("  %-16s %4d 条指令" % (c, cats[c]))


def cmd_check(args):
    csv_path = Path(args.csv)
    ref_path = Path(args.ref)

    if not csv_path.exists():
        print("错误: 找不到 CSV 文件: %s" % csv_path, file=sys.stderr)
        sys.exit(1)

    # 加载参考值 —— 优先用 JSON，否则直接读 Excel
    if ref_path.exists():
        ref = load_reference(ref_path)
        print(dim("参考值: %s (%d 条指令)" % (ref_path, len(ref))))
    elif DEFAULT_XLSX.exists():
        print(dim("参考 JSON 不存在，直接读取 Excel: %s" % DEFAULT_XLSX))
        ref = parse_excel(DEFAULT_XLSX)
    else:
        print("错误: 找不到参考数据，请先执行 'uarch update'。", file=sys.stderr)
        sys.exit(1)

    print(dim("结果文件: %s" % csv_path))
    print(dim("比对方式: 四舍五入取整后对比"))

    matches, mismatches, skipped = check_results(csv_path, ref)
    ok = print_report(matches, mismatches, skipped)

    sys.exit(0 if ok else 1)


def main():
    p = argparse.ArgumentParser(prog="refcheck",
                                description="uarch-bench 参考值比对工具")
    sub = p.add_subparsers(dest="command")

    # update
    up = sub.add_parser("update", help="从 Excel 提取参考值 → JSON")
    up.add_argument("--xlsx", default=str(DEFAULT_XLSX),
                    help="Excel 文件路径 (默认: doc/RVA23_Instruction_Classification.xlsx)")
    up.add_argument("--out", default=str(DEFAULT_REF),
                    help="输出 JSON 路径 (默认: data/reference.json)")

    # check
    ch = sub.add_parser("check", help="比对 CSV 结果与参考值")
    ch.add_argument("csv", help="results.csv 路径")
    ch.add_argument("--ref", default=str(DEFAULT_REF),
                    help="参考 JSON 路径 (默认: data/reference.json)")

    args = p.parse_args()
    if args.command == "update":
        cmd_update(args)
    elif args.command == "check":
        cmd_check(args)
    else:
        p.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
